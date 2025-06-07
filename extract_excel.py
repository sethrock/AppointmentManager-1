#!/usr/bin/env python3
import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('attached_assets/Client names for testing Google Calendar_1749265665332.xlsx')
ws = wb.active

# Extract data
data = []
headers = []

# Get headers from first row
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# Get data from subsequent rows
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(row):  # Skip empty rows
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = value
        data.append(row_data)

# Save as JSON for easier processing
with open('test_calendar_data.json', 'w') as f:
    json.dump(data, f, indent=2, default=str)

print(f"Extracted {len(data)} rows of test data")
print("Headers:", headers)
print("Sample data:", data[:3] if data else "No data")